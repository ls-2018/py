from openpyxl import Workbook

wb = Workbook()  # 默认有一个Sheet工作薄

# 创作一个新的工作薄
ws1 = wb.create_sheet('Mysheet')  # 在最后添加
ws2 = wb.create_sheet('Mysheeta', 0)  # 在第一个添加

# 修改工作薄的名称
ws2.title = "New Title"

# 查看工作薄的名称列表        ['New Title', 'Sheet', 'Mysheet']
print(wb.sheetnames)

# 写数据
ws2["A4"] = 4
ws2.cell(row=4, column=2, value=10)  # 按单元格来区分,从1开始

l = [1, 2, 3, 4, '', 6]
ws2.append(l)  # 在最下边,添加了一行

# Excel内置函数
"""
ws2['G5'] = '=SUM(A5:F5)'       #  一片区域:=SUM(A6:F10)
"""
# 保存
wb.save('balances.xlsx')
