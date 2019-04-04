from openpyxl import load_workbook

wb = load_workbook('balances.xlsx', read_only=False, data_only=True)
# 控制带有公式的单元格是否具有公式(默认值)   Excel上次读取工作表时存储的值

print(wb.sheetnames)  # ['New Title', 'Sheet', 'Mysheet']
wb1 = wb['New Title']
# ws = wb.active     #获取第一个sheet
# wb.get_sheet_by_name("New Title"  )

# wb1.guess_types = False
# wb1["D1"]="12%"
# wb1.guess_types = True    print ws["D1"].value会打印百分数b
# wb1.guess_types = False   print ws["D1"].value会打印小数

print(wb1['B4'].value)
print(wb1.cell(row=4, column=2).value)

print(wb1.max_row)  # 最大行数
print(wb1.max_column)  # 最大列数

# ######################## 获取行和列 ####################################
# 因为按行，所以返回A1, B1, C1这样的顺序
for row in wb1.rows:
    for cell in row:
        print(cell.value)

# A1, A2, A3这样的顺序
# for column in wb1.columns:    # read_only=False
#     for cell in column:
#         print(cell.value)

for row_cell in wb1['A1':'B3']:
    for cell in row_cell:
        print(cell.value)

# ######################## 设置字体 ####################################
from openpyxl.styles import Font, colors

bold_itatic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED, bold=True, underline="single")
wb1['A1'].font = bold_itatic_24_font
wb1['A1'] = 'SX'

# ######################## 对齐方式 ####################################
from openpyxl.styles import Alignment

for row_cell in wb1['B2':'G10']:
    for cell in row_cell:
        cell.alignment = Alignment(horizontal='center', vertical='center')  # right,left

# ######################## 设置行高和列宽 ####################################
# 有时候数据太长显示不完，就需要拉长拉高单元格
# 第2行行高
wb1.row_dimensions[2].height = 40
# C列列宽
wb1.column_dimensions['C'].width = 30

# ######################## 合并和拆分单元格 ####################################
wb1.merge_cells('B1:G1')  # 合并一行中的几个单元格
"""
合并后只需要往第一个表格写数据
如果这些要合并的单元格都有数据，只会保留左上角的数据，其他则丢弃。
换句话说若合并前不是在左上角写入数据，合并后单元格中不会有数据。
"""
# wb1.unmerge_cells('A1:C3')


# ############# 设定单元格的边框、字体、颜色、大小和边框背景色 ###############
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill

highlight = NamedStyle(name="highlight")
highlight.font = Font(name='微软雅黑', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False,
                      color='FF000000')
highlight.fill = PatternFill("solid", fgColor="DDDDDD")  # 背景填充
bd = Side(style='thick', color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)

print(dir(wb1["A1"]))
wb1["A1"].style = highlight
# ############# 常用的样式和属性设置 ###############


from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

fill = PatternFill(fill_type="solid", start_color='FFEEFFFF', end_color='FF001100')

# 边框可以选择的值为：'hair', 'medium', 'dashDot', 'dotted', 'mediumDashDot', 'dashed', 'mediumDashed', 'mediumDashDotDot', 'dashDotDot', 'slantDashDot', 'double', 'thick', 'thin']
# diagonal 表示对角线
bd = Border(left=Side(border_style="thin", color='FF001000'),
            right=Side(border_style="thin", color='FF110000'),
            top=Side(border_style="thin", color='FF110000'),
            bottom=Side(border_style="thin", color='FF110000'),
            diagonal=Side(border_style=None, color='FF000000'),
            diagonal_direction=0, outline=Side(border_style=None, color='FF000000'),
            vertical=Side(border_style=None, color='FF000000'),
            horizontal=Side(border_style=None, color='FF110000'))

alignment = Alignment(horizontal='general', vertical='bottom', text_rotation=0, wrap_text=False, shrink_to_fit=False,
                      indent=0)
number_format = 'General'
protection = Protection(locked=True, hidden=False)

wb1["B5"].fill = fill
wb1["B5"].border = bd
wb1["B5"].alignment = alignment
wb1["B5"].number_format = number_format

wb.save('balances.xlsx')
