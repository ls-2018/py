# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series

wb = load_workbook('./sample.xlsx')
ws1=wb.active

wb = Workbook()
ws = wb.active
for i in range(10):
    ws.append([i])

values = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=10)
chart = BarChart()
chart.add_data(values)
ws.add_chart(chart, "E15")

# Save the file
wb.save("./sample1.xlsx")